from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl as xl
import pandas as pd
import os
from datetime import date
import numpy as np

class Spreadsheet:
    def __init__(self, wb):
        self.wb = wb
        self.sheets = ["CoverSheet", "Work Profiles", "Loads", "Cylinder", "Statistics", "Load Severity", "Histogram Charts", "Lift Pressure Histograms", "Tilt Pressure Histograms", "Steering Pressure Histograms", "Force Histograms"]
        self.filename = "test3-2024-07-25.xlsx"
        self.hardbank = np.array([['Platform', '950 962 M&L', '950 962 AU2020', '966 972', '980 982'],
                            ['Hardbank', 70, 45, 90, 90],
                            ['Truck Loading', 0, 25, 0, 0],
                            ['Bulldoze/Backdrag', 9, 9, 8, 8],
                            ['Non-Damaging', 21, 21, 2, 2],
                            ['Sum', 100, 100, 100, 100]])
        
        self.truck_loading =  np.array([['Platform', '950 962 M&L', '950 962 AU2020', '966 972', '980 982'],
                            ['Hardbank', 10, 10, 0, 0],
                            ['Truck Loading–2" Rock', 35, 35, 45, 45],
                            ['Truck Loading–Pea Gravel', 35, 35, 45, 45],
                            ['Bulldoze/Backdrag', 9, 9, 8, 8],
                            ['Non-Damaging', 11, 11, 2, 2],
                            ['Sum', 100, 100, 100, 100]])

#Constant sheets:
    def construct(self, path):
        def get_last_folder(path):
            return os.path.basename(os.path.normpath(path))

        folder = get_last_folder(path)
        for i in range(1, len(self.sheets)):
            self.wb.create_sheet(self.sheets[i])
        today = date.today()
        def get_folder_after_users(path):
            folders = path.split(os.sep)
            try:
                users_index = folders.index('Users')
                return folders[users_index + 1]
            except ValueError:
                return "The path does not contain a 'Users' directory."
        engineer = get_folder_after_users(path)
        
        ws1 = self.wb["Sheet"]
        ws1.title = "CoverSheet"
        ws1["A1"] = "PROGRAM: MWL Hydraulic Systems - Cylinders"
        ws1["A3"] = f"PROJECT: {folder} Supplier Designed Cylinders"
        ws1["A5"] = "DESCRIPTION: Cylinder Load & Load Severity Calculations"
        ws1["A7"] = f"ENGINEER: {engineer}"
        ws1["A9"] = f"DATE: {today}"
        ws1["A11"] = f"DATA LOCATION: {path}"
        ws1["A13"] = "NOTES: "
        ws1["A15"] = "ABSTRACT:"

        return folder

        self.wb.save(self.filename)

    def constant_sheet(self, filepath, sheet):
        path1 = filepath
        wb1 = xl.load_workbook(filename=path1)
        ws1 = wb1.worksheets[0]

        ws = self.wb[sheet]
        for row in ws1:
            for cell in row:
                ws[cell.coordinate].value = cell.value

        self.wb.save(self.filename)
     
    def average_pressure_loads(self):

        cylinders = ["Lift Pressure Histograms", "Steering Pressure Histograms"]
        ws1 = self.wb[cylinders[0]]
        ws2 = self.wb[cylinders[1]]
        wss = [ws1, ws2]
        ws = self.wb["Loads"]

        scenes = ["700", "710", "720", "730", "741", "757", "760"]

        lift_counters = [0, 0, 0, 0, 0, 0, 0]
        column_counter = 0
        histogram_index = 1
        
        for i in range(len(cylinders)): 
            for j in range(1,len(wss[i][1]), 5):
                column_counter += 1
                for k in range(len(scenes)): 
                    if self.contains_substring(wss[i].cell(row=1, column=j).value, scenes[k]) and lift_counters[k] < 2:
                        current_index = j
                        for l in range(2,25):
                            avg = (wss[i].cell(row=l, column=current_index+3).value + wss[i].cell(row=l, column=current_index+13).value)/2
                            ws.cell(row=l+histogram_index, column=column_counter, value=avg)
                        ws.cell(row=histogram_index, column=column_counter, value=wss[i].cell(row=1, column=current_index+1).value)
                        ws.cell(row=histogram_index+1, column=column_counter, value=wss[i].cell(row=1, column=current_index).value)
                        lift_counters[k] += 1       
            self.reset_lift_counters(lift_counters)
            histogram_index += 26
            column_counter = 0

        self.wb.save(self.filename)

    def reset_lift_counters(self, lift_counters):
        for i in range(len(lift_counters)):
            lift_counters[i] = 0

    def contains_substring(self, main_string, substring):
        return substring in main_string


    def statistics(self, results, num, filename, cylinder):
        df = pd.DataFrame(results)

        #load workbook
        ws = self.wb['Statistics']
        rows = dataframe_to_rows(df, index=False)
        for r_idx, row in enumerate(rows, 1):
            ws.cell(row=r_idx+num*3, column=1, value=filename)
            ws.cell(row=r_idx+num*3, column=2, value=cylinder)
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx+num*3, column=c_idx+2, value=value)

        self.wb.save(self.filename)

    def fela(self, results, num):
        df = pd.DataFrame(results)

        #load workbook
        ws = self.wb['Loads']
        rows = dataframe_to_rows(df, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx+num*3+62, column=c_idx, value=value)

        self.wb.save(self.filename)

    def load_severity(self, load_severity_value, current_file, combine, func, num_stat, sn_curve):
        ws = self.wb['Load Severity']

        ws.cell(row=1, column=1, value="File") 
        ws.cell(row=1, column=2, value="Cylinder") 
        ws.cell(row=1, column=3, value="Function") 
        ws.cell(row=1, column=4, value="SN-Curve") 
        ws.cell(row=1, column=5, value="Load Severity") 
        ws.cell(row=1, column=1+5, value="File") 
        ws.cell(row=1, column=2+5, value="Cylinder") 
        ws.cell(row=1, column=3+5, value="Function") 
        ws.cell(row=1, column=4+5, value="SN-Curve") 
        ws.cell(row=1,column=5+5, value="Load Severity") 
            

        if sn_curve == 3:
            ws.cell(row=num_stat+1, column=1, value=current_file) 
            ws.cell(row=num_stat+1, column=2, value=combine)
            ws.cell(row=num_stat+1, column=3, value=func)
            ws.cell(row=num_stat+1, column=4, value=sn_curve)
            ws.cell(row=num_stat+1, column=5, value=load_severity_value)
        else:
            ws.cell(row=num_stat+1, column=1+5, value=current_file) 
            ws.cell(row=num_stat+1, column=2+5, value=combine)
            ws.cell(row=num_stat+1, column=3+5, value=func)
            ws.cell(row=num_stat+1, column=4+5, value=sn_curve)
            ws.cell(row=num_stat+1, column=5+5, value=load_severity_value)
 

    def composite_load_severity(self):
        ws = self.wb['Load Severity']
        ws2 = self.wb['Loads']



        ws2.cell(row=55, column=1, value="Lift Head Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=55, column=2, value="Lift Rod Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=55, column=3, value="Lift Head Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=55, column=4, value="Lift Rod Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=55, column=5, value="Tilt Head Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=55, column=6, value="Tilt Rod Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=55, column=7, value="Tilt Head Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=55, column=8, value="Tilt Rod Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=55, column=9, value="Steering Head Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=55,column=10, value="Steering Rod Hardbank Total Load Severity, SN=3")
        ws2.cell(row=55, column=11, value="Steering Head Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=55,column=12, value="Steering Rod Truck Loading Total Load Severity, SN=3")

        ws2.cell(row=56, column=1, value=self.avg(ws.cell(row=52, column=5).value, ws.cell(row=54, column=5).value)*.7+self.avg(ws.cell(row=62, column=5).value, ws.cell(row=64, column=5).value)*.09)
        ws2.cell(row=56, column=2, value=self.avg(ws.cell(row=53, column=5).value, ws.cell(row=55, column=5).value)*.7+self.avg(ws.cell(row=63, column=5).value, ws.cell(row=65, column=5).value)*.09)
        ws2.cell(row=56, column=3, value=self.avg(ws.cell(row=2, column=5).value, ws.cell(row=4, column=5).value)*.35+self.avg(ws.cell(row=12, column=5).value, ws.cell(row=14, column=5).value)*.35+self.avg(ws.cell(row=52, column=5).value, ws.cell(row=54, column=5).value)*.1+self.avg(ws.cell(row=62, column=5).value, ws.cell(row=64, column=5).value)*.09)
        ws2.cell(row=56, column=4, value=self.avg(ws.cell(row=3, column=5).value, ws.cell(row=5, column=5).value)*.35+self.avg(ws.cell(row=13, column=5).value, ws.cell(row=15, column=5).value)*.35+self.avg(ws.cell(row=53, column=5).value, ws.cell(row=55, column=5).value)*.1+self.avg(ws.cell(row=63, column=5).value, ws.cell(row=65, column=5).value)*.09)

        ws2.cell(row=56, column=5, value=ws.cell(row=60, column=5).value*.7+ws.cell(row=70, column=5).value*.09)
        ws2.cell(row=56, column=6, value=ws.cell(row=61, column=5).value*.7+ws.cell(row=71, column=5).value*.09)
        ws2.cell(row=56, column=7, value=ws.cell(row=10, column=5).value*.35+ws.cell(row=20, column=5).value*.35+ws.cell(row=60, column=5).value*.1+ws.cell(row=70, column=5).value*.09)
        ws2.cell(row=56, column=8, value=ws.cell(row=11, column=5).value*.35+ws.cell(row=21, column=5).value*.35+ws.cell(row=61, column=5).value*.1+ws.cell(row=71, column=5).value*.09)

        ws2.cell(row=56, column=9, value=self.avg(ws.cell(row=56, column=5).value, ws.cell(row=58, column=5).value)*.7+self.avg(ws.cell(row=66, column=5).value, ws.cell(row=68, column=5).value)*.09)
        ws2.cell(row=56, column=10, value=self.avg(ws.cell(row=53, column=5).value, ws.cell(row=55, column=5).value)*.7+self.avg(ws.cell(row=67, column=5).value, ws.cell(row=69, column=5).value)*.09)
        ws2.cell(row=56, column=11, value=self.avg(ws.cell(row=6, column=5).value, ws.cell(row=8, column=5).value)*.35+self.avg(ws.cell(row=16, column=5).value, ws.cell(row=18, column=5).value)*.35+self.avg(ws.cell(row=56, column=5).value, ws.cell(row=58, column=5).value)*.1+self.avg(ws.cell(row=66, column=5).value, ws.cell(row=68, column=5).value)*.09)
        ws2.cell(row=56, column=12, value=self.avg(ws.cell(row=7, column=5).value, ws.cell(row=9, column=5).value)*.35+self.avg(ws.cell(row=17, column=5).value, ws.cell(row=19, column=5).value)*.35+self.avg(ws.cell(row=57, column=5).value, ws.cell(row=59, column=5).value)*.1+self.avg(ws.cell(row=67, column=5).value, ws.cell(row=69, column=5).value)*.09)


        ws2.cell(row=58, column=1, value="Lift Head Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=58, column=2, value="Lift Rod Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=58, column=3, value="Lift Head Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=58, column=4, value="Lift Rod Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=58, column=5, value="Tilt Head Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=58, column=6, value="Tilt Rod Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=58, column=7, value="Tilt Head Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=58, column=8, value="Tilt Rod Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=58, column=9, value="Steering Head Hardbank Total Load Severity, SN=3") 
        ws2.cell(row=58,column=10, value="Steering Rod Hardbank Total Load Severity, SN=3")
        ws2.cell(row=58, column=11, value="Steering Head Truck Loading Total Load Severity, SN=3") 
        ws2.cell(row=58,column=12, value="Steering Rod Truck Loading Total Load Severity, SN=3")

        ws2.cell(row=59, column=1, value=self.avg(ws.cell(row=52, column=10).value, ws.cell(row=54, column=10).value)*.7+self.avg(ws.cell(row=62, column=10).value, ws.cell(row=64, column=10).value)*.09)
        ws2.cell(row=59, column=2, value=self.avg(ws.cell(row=53, column=10).value, ws.cell(row=55, column=10).value)*.7+self.avg(ws.cell(row=63, column=10).value, ws.cell(row=65, column=10).value)*.09)
        ws2.cell(row=59, column=3, value=self.avg(ws.cell(row=2, column=10).value, ws.cell(row=4, column=10).value)*.35+self.avg(ws.cell(row=12, column=10).value, ws.cell(row=14, column=10).value)*.35+self.avg(ws.cell(row=52, column=10).value, ws.cell(row=54, column=10).value)*.1+self.avg(ws.cell(row=62, column=10).value, ws.cell(row=64, column=10).value)*.09)
        ws2.cell(row=59, column=4, value=self.avg(ws.cell(row=3, column=10).value, ws.cell(row=5, column=10).value)*.35+self.avg(ws.cell(row=13, column=10).value, ws.cell(row=15, column=10).value)*.35+self.avg(ws.cell(row=53, column=10).value, ws.cell(row=55, column=10).value)*.1+self.avg(ws.cell(row=63, column=10).value, ws.cell(row=65, column=10).value)*.09)

        ws2.cell(row=59, column=5, value=ws.cell(row=60, column=10).value*.7+ws.cell(row=70, column=10).value*.09)
        ws2.cell(row=59, column=6, value=ws.cell(row=61, column=10).value*.7+ws.cell(row=71, column=10).value*.09)
        ws2.cell(row=59, column=7, value=ws.cell(row=10, column=10).value*.35+ws.cell(row=20, column=10).value*.35+ws.cell(row=60, column=10).value*.1+ws.cell(row=70, column=10).value*.09)
        ws2.cell(row=59, column=8, value=ws.cell(row=11, column=10).value*.35+ws.cell(row=21, column=10).value*.35+ws.cell(row=61, column=10).value*.1+ws.cell(row=71, column=10).value*.09)

        ws2.cell(row=59, column=9, value=self.avg(ws.cell(row=56, column=10).value, ws.cell(row=58, column=10).value)*.7+self.avg(ws.cell(row=66, column=10).value, ws.cell(row=68, column=10).value)*.09)
        ws2.cell(row=59, column=10, value=self.avg(ws.cell(row=53, column=10).value, ws.cell(row=55, column=10).value)*.7+self.avg(ws.cell(row=67, column=10).value, ws.cell(row=69, column=10).value)*.09)
        ws2.cell(row=59, column=11, value=self.avg(ws.cell(row=6, column=10).value, ws.cell(row=8, column=10).value)*.35+self.avg(ws.cell(row=16, column=10).value, ws.cell(row=18, column=10).value)*.35+self.avg(ws.cell(row=56, column=10).value, ws.cell(row=58, column=10).value)*.1+self.avg(ws.cell(row=66, column=10).value, ws.cell(row=68, column=10).value)*.09)
        ws2.cell(row=59, column=12, value=self.avg(ws.cell(row=7, column=10).value, ws.cell(row=9, column=10).value)*.35+self.avg(ws.cell(row=17, column=10).value, ws.cell(row=19, column=10).value)*.35+self.avg(ws.cell(row=57, column=10).value, ws.cell(row=59, column=10).value)*.1+self.avg(ws.cell(row=67, column=10).value, ws.cell(row=69, column=10).value)*.09)
 

        self.wb.save(self.filename)

    def avg(self, a, b):
        return (a+b)/2


    def histogram_chart(self, image_path, num):
        ws = self.wb["Histogram Charts"]
        img = Image(image_path)
        pos = 'A'+str(num)
        img.anchor = pos
        ws.add_image(img)
        self.wb.save(self.filename)

    def pressure_histograms(self, results, num, filename, cylinder, func):
        df = results
        sheets = ['Lift Pressure Histograms', 'Tilt Pressure Histograms', 'Steering Pressure Histograms']
        if func == 'LFT': 
            ws = self.wb[sheets[0]]
        elif func == 'TLT':
            ws = self.wb[sheets[1]]
        else:
            ws = self.wb[sheets[2]]

        rows = dataframe_to_rows(df, index=False)
        for r_idx, row in enumerate(rows, 1):
            ws.cell(row=r_idx, column=1+num*5, value=filename)
            ws.cell(row=r_idx, column=2+num*5, value=cylinder)
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx+2+num*5, value=value)

        self.wb.save(self.filename)

    def force_histograms(self):
        pass

    def composite_lift_head(self):
        ws = self.wb['Loads']
        ws2 = self.wb['Histogram Charts']
        ws3 = self.wb['Lift Pressure Histograms']
        data1 = []
        data2 = []
        data = [data1, data2]

        ws2.cell(row=1, column=29, value="Lift Cylinder Head End Hardbank")
        for j in range(3, 26):
            composite = (ws.cell(row=j, column=21).value)*0.7+(ws.cell(row=j, column=25).value)*.09
            ws2.cell(row=j-1, column=29, value=composite)
            data1.append(composite)
        self.wb.save(self.filename)
        for j in range(2, 25):
            ws2.cell(row=j, column=28, value=ws3.cell(row=j, column=3).value)
 
        ws2.cell(row=1, column=30, value="Lift Cylinder Head End TruckLoading")
        for j in range(3, 26):
            composite = (ws.cell(row=j, column=21).value)*0.1+(ws.cell(row=j, column=1).value)*.35+(ws.cell(row=j, column=5).value)*.35+(ws.cell(row=j, column=25).value)*.09
            ws2.cell(row=j-1, column=30, value=composite)
            data2.append(composite)
        self.wb.save(self.filename)
        return data
    
    def composite_lift_rod(self):
        ws = self.wb['Loads']
        ws2 = self.wb['Histogram Charts']
        ws3 = self.wb['Lift Pressure Histograms']
        data1 = []
        data2 = []
        data = [data1, data2]

        ws2.cell(row=1, column=31, value="Lift Cylinder Rod End Hardbank")
        for j in range(3, 26):
            composite = (ws.cell(row=j, column=22).value)*0.7+(ws.cell(row=j, column=26).value)*.09
            ws2.cell(row=j-1, column=32, value=composite)
            data1.append(composite)
        self.wb.save(self.filename)
        for j in range(2, 25):
            ws2.cell(row=j, column=31, value=ws3.cell(row=j, column=3).value)

        ws2.cell(row=1, column=32, value="Lift Cylinder Rod End TruckLoading")
        for j in range(3, 26):
            composite = (ws.cell(row=j, column=22).value)*0.1+(ws.cell(row=j, column=2).value)*.35+(ws.cell(row=j, column=6).value)*.35+(ws.cell(row=j, column=26).value)*.09
            ws2.cell(row=j-1, column=33, value=composite)
            data2.append(composite)
        self.wb.save(self.filename)
        return data
    
    def composite_steer_head(self):
        ws = self.wb['Loads']
        ws2 = self.wb['Histogram Charts']
        ws3 = self.wb['Steering Pressure Histograms']
        data1 = []
        data2 = []
        data = [data1, data2]

        ws2.cell(row=1+25, column=29, value="Steer Cylinder Head End Hardbank")
        for j in range(3+26, 26+26):
            composite = (ws.cell(row=j, column=21).value)*0.7+(ws.cell(row=j, column=25).value)*.09
            ws2.cell(row=j-1, column=29, value=composite)
            data1.append(composite)
        self.wb.save(self.filename)
        for j in range(2, 25):
            ws2.cell(row=j+26, column=28, value=ws3.cell(row=j, column=3).value)

        ws2.cell(row=1+25, column=30, value="Lift Cylinder Head End TruckLoading")
        for j in range(3+26, 26+26):
            composite = (ws.cell(row=j, column=21).value)*0.1+(ws.cell(row=j, column=1).value)*.35+(ws.cell(row=j, column=5).value)*.35+(ws.cell(row=j, column=25).value)*.09
            ws2.cell(row=j-1, column=30, value=composite)
            data2.append(composite)
        self.wb.save(self.filename)
        return data
    
    def composite_steer_rod(self):
        ws = self.wb['Loads']
        ws2 = self.wb['Histogram Charts']
        ws3 = self.wb['Steering Pressure Histograms']
        data1 = []
        data2 = []
        data = [data1, data2]

        ws2.cell(row=1+25, column=31, value="Steer Cylinder Rod End Hardbank")
        for j in range(3+26, 26+26):
            composite = (ws.cell(row=j, column=22).value)*0.7+(ws.cell(row=j, column=26).value)*.09
            ws2.cell(row=j-1, column=32, value=composite)
            data1.append(composite)
        self.wb.save(self.filename)
        for j in range(2, 25):
            ws2.cell(row=j+26, column=31, value=ws3.cell(row=j, column=3).value)

        ws2.cell(row=1+25, column=32, value="Steer Cylinder Rod End TruckLoading")
        for j in range(3+26, 26+26):
            composite = (ws.cell(row=j, column=22).value)*0.1+(ws.cell(row=j, column=2).value)*.35+(ws.cell(row=j, column=6).value)*.35+(ws.cell(row=j, column=26).value)*.09
            ws2.cell(row=j-1, column=33, value=composite)
            data2.append(composite)
        self.wb.save(self.filename)
        return data
    
    def composite_tilt_head(self):
        ws = self.wb['Tilt Pressure Histograms']
        ws2 = self.wb['Histogram Charts']
        ws3 = self.wb['Tilt Pressure Histograms']
        data1 = []
        data2 = []
        data = [data1, data2]

        ws2.cell(row=1+25+25, column=29, value="Tilt Cylinder Head End Hardbank")
        for j in range(2, 27):
            composite = (ws.cell(row=j, column=54).value)*0.7+(ws.cell(row=j, column=64).value)*.09
            ws2.cell(row=j-1+52, column=29, value=composite)
            data1.append(composite)
        self.wb.save(self.filename)
        for j in range(2, 25):
            ws2.cell(row=j+52, column=28, value=ws3.cell(row=j, column=3).value)

        ws2.cell(row=1+25+25, column=30, value="Tilt Cylinder Head End TruckLoading")
        for j in range(2, 27):
            composite = (ws.cell(row=j, column=54).value)*0.1+(ws.cell(row=j, column=4).value)*.35+(ws.cell(row=j, column=14).value)*.35+(ws.cell(row=j, column=64).value)*.09
            ws2.cell(row=j-1+52, column=30, value=composite)
            data2.append(composite)
        self.wb.save(self.filename)
        return data
    
    def composite_tilt_rod(self):
        ws = self.wb['Tilt Pressure Histograms']
        ws2 = self.wb['Histogram Charts']
        ws3 = self.wb['Tilt Pressure Histograms']
        data1 = []
        data2 = []
        data = [data1, data2]

        ws2.cell(row=1+25+25, column=31, value="Tilt Cylinder Rod End Hardbank")
        for j in range(2, 27):
            composite = (ws.cell(row=j, column=54+5).value)*0.7+(ws.cell(row=j, column=64+5).value)*.09
            ws2.cell(row=j-1, column=32, value=composite)
            data1.append(composite)
        self.wb.save(self.filename)
        for j in range(2, 25):
            ws2.cell(row=j+52, column=31, value=ws3.cell(row=j, column=3).value)

        ws2.cell(row=+25+25, column=32, value="Tilt Cylinder Rod End TruckLoading")
        for j in range(2, 27):
            composite = (ws.cell(row=j, column=54+5).value)*0.1+(ws.cell(row=j, column=4+5).value)*.35+(ws.cell(row=j, column=14+5).value)*.35+(ws.cell(row=j, column=64+5).value)*.09
            ws2.cell(row=j-1, column=33, value=composite)
            data2.append(composite)
        self.wb.save(self.filename)
        return data
    



    def get_column(self):
        return self.hardbank[:, 1]